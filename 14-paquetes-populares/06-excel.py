import openpyxl

# cuando puse solo el nombre del archivo no lo encontro dentro de la misma carpeta asi que tuve que indicar la ruta
# o la otra forma es tener el archivo en la raiz
# load_workbook es para cargar un excel existente
wb = openpyxl.load_workbook("14-paquetes-populares/winrate.xlsx")

print('nombre hojas', wb.sheetnames) # nombres de las hojas que tenga el xlsx
ws = wb["Hoja 1"]  # seleccionamos una sola hoja del workbook, si no se especifica toma por defecto la
print('hoja 1', ws)
# otra forma seria
hoja = wb.active
print('hoja activa',hoja)

# agregar una nueva hoja
wb.create_sheet("hoja creada")
hoja2 = wb["hoja creada"]
hoja2.title = "Nuevo titulo en hoja creada"
print(wb.sheetnames) # nombres de las hojas

# print(
#     wb.max_row,
#     wb.max_column,
# )

celda = ws["A2"]
print('valor celda',celda.value)
# cambiar valor de la celda
celda.value = "Nombre modificado"
print('valor celda modificado',celda.value)



# otro metodo de aceder a distintas celdas
celda2 = ws.cell(row=2, column=1)
print(
    'celda2',
    celda2.value, # Nombre modificado
    celda2.row, # 2
    celda2.column, # 1
    celda2.coordinate #A2
)

# podemos acceder a los valores de las filas y columnas de la siguiente manera
for fila in range(1, ws.max_row + 1):
    for columna in range(1, ws.max_column + 1):
        celda = ws.cell(row=fila, column=columna)
        print('iteracion excel', fila, columna, celda.value)

# otra forma de este mismo ejemplo seria el siguiente aunque su orden de recorrido es distinto
for fila in ws.iter_rows():
    for celda in fila:
        valor = celda.value
        # Aquí puedes trabajar con el valor de la celda como desees
        print('iter_rows', valor)


# obtener una columna o fila en particular
columna = ws["A"]
fila = ws["2"]
print('columna completa', columna)
print('fila completa', fila)
#  y acceder a sus valores seria por ejemplo de la siguiente manera
for celda in columna:
    valor = celda.value
    print('solo columna A',valor)

for celda in fila:
    valor = celda.value
    print('solo fila 2',valor)


# agregar mas elementos a la fila
ws.append([1,2,3])

# eliminar columnas y filas
ws.delete_cols(1,1) # indice desde donde quiero iniciara eliminar y cantidad de filas o columns si no se entrega valor por defecto sera 1
ws.delete_rows(1,2)

# crear un nuevo excel con las modificaciones
wb.save("14-paquetes-populares/nuevo_excel.xlsx") #nombre del archivo y ruta
